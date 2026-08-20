import csv
import io
from django.http import HttpResponse
from django.template.loader import render_to_string

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def export_csv(filename, headers, rows):
    # charset must be plain utf-8: Django encodes EVERY response.write() with
    # the declared codec, so 'utf-8-sig' prepended a BOM to every row and
    # corrupted the first column of every line. Write the BOM once instead —
    # Excel still needs it to detect UTF-8.
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    response.write('﻿')
    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow([str(v) if v is not None else '' for v in row])
    return response


def render_report_pdf(request, template, ctx, filename, landscape=True):
    """Render a print template to a downloadable PDF (Phase 17.2 / 17.6).

    Engine cascade: WeasyPrint (correct Bangla shaping — used on the Linux/Docker
    server) → xhtml2pdf (pure-Python fallback for Windows dev; Bangla shaping is
    limited but it always produces a valid PDF). Returns the PDF as an
    ``attachment`` so the browser downloads it instead of opening a print dialog.
    """
    from pathlib import Path
    from django.conf import settings

    ctx = dict(ctx)
    ctx['pdf_mode']      = True          # suppress the auto window.print() script
    ctx['pdf_landscape'] = landscape     # A4 landscape @page for wide tables

    # Embed the Bangla font by absolute file URI so WeasyPrint can load it
    # without depending on collected static files.
    font_path = Path(settings.BASE_DIR) / 'static' / 'fonts' / 'SolaimanLipi.ttf'
    ctx['pdf_font_uri'] = font_path.as_uri() if font_path.exists() else ''

    html = render_to_string(template, ctx, request=request)

    pdf_bytes = None
    try:
        from weasyprint import HTML as WP_HTML
        pdf_bytes = WP_HTML(string=html, base_url=str(settings.BASE_DIR)).write_pdf()
    except Exception:
        pdf_bytes = None

    if pdf_bytes is None:
        from xhtml2pdf import pisa
        buf = io.BytesIO()
        pisa.pisaDocument(io.BytesIO(html.encode('utf-8')), buf)
        pdf_bytes = buf.getvalue()

    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def export_excel(filename, headers, rows, sheet_title='রিপোর্ট'):
    if not HAS_OPENPYXL:
        return HttpResponse('openpyxl not installed', status=500)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill('solid', fgColor='0D9488')
    header_font = Font(bold=True, color='FFFFFF')
    alt_fill = PatternFill('solid', fgColor='F0FDFA')

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 28

    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val is not None else '')
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    for col in ws.columns:
        max_len = max((len(str(c.value or '')) for c in col if c.value), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response
