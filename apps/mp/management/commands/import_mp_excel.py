import os
from datetime import datetime

import openpyxl
from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.master.models import District, Division, Gender, Religion
from apps.mp.models import MP, Address, ElectionInfo
from apps.parliament.models import Constituency, Parliament

User = get_user_model()

# ── Alias maps (Excel value → DB name_en, all lower-case keys) ────────────────

DIVISION_ALIASES = {
    'barisal':    'barishal',
    'chittagong': 'chattogram',
}

DISTRICT_ALIASES = {
    'bogra':             'bogura',
    'chapai nababganj':  'chapainawabganj',
    'jessore':           'jashore',
    'barisal':           'barishal',
    'jhalokati':         'jhalakathi',
    'netrakona':         'netrokona',
    'kishoregonj':       'kishoreganj',
    'comilla':           'cumilla',
    'chittagong':        'chattogram',
    "cox's bazar":       'coxsbazar',
    'khagrachhari':      'khagrachari',
}

RELIGION_ALIASES = {
    'hinduism':  'shonaton (hinduism)',
    'buddhisn':  'buddhism',
}

MEMBER_TYPE_MAP = {
    'directly elected':      'direct',
    'reserved seat (women)': 'reserved',
}


def _norm_mobile(raw):
    """Convert 880XXXXXXXXXX (13-digit) → 01XXXXXXXXX (11-digit)."""
    s = str(raw).strip() if raw else ''
    if s.startswith('880') and len(s) == 13:
        return '0' + s[3:]
    return s


def _parse_date(raw):
    """Parse DD/MM/YYYY → date object, or return None."""
    if not raw:
        return None
    try:
        return datetime.strptime(str(raw).strip(), '%d/%m/%Y').date()
    except ValueError:
        return None


def _resolve(lookup_map, raw_value, alias_map=None):
    """Case-insensitive FK lookup with optional alias map. Returns object or None."""
    if not raw_value:
        return None
    key = str(raw_value).strip().lower()
    if alias_map:
        key = alias_map.get(key, key)
    return lookup_map.get(key)


class Command(BaseCommand):
    help = 'Import MP data from Excel file (docs/mp_report.xlsx)'

    def add_arguments(self, parser):
        parser.add_argument(
            'excel_file',
            nargs='?',
            default='docs/mp_report.xlsx',
            help='Path to the Excel file (default: docs/mp_report.xlsx)',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Validate and report without saving anything to the database',
        )
        parser.add_argument(
            '--skip-existing',
            action='store_true',
            default=True,
            help='Skip rows where mp_id already exists (default: True)',
        )

    def handle(self, *args, **options):
        excel_path = options['excel_file']
        dry_run    = options['dry_run']
        skip_existing = options['skip_existing']

        if not os.path.exists(excel_path):
            raise CommandError(f'File not found: {excel_path}')

        self.stdout.write(f'Reading: {excel_path}')
        if dry_run:
            self.stdout.write(self.style.WARNING('DRY RUN — no data will be saved'))

        # ── Pre-load all lookup maps ──────────────────────────────────────────
        parliament = Parliament.objects.filter(name_en='13th Parliament').first()
        if not parliament:
            raise CommandError('Parliament "13th Parliament" not found in database')

        religion_map     = {r.name_en.lower(): r for r in Religion.objects.all()}
        district_map     = {d.name_en.lower(): d for d in District.objects.all()}
        division_map     = {d.name_en.lower(): d for d in Division.objects.all()}
        gender_map       = {g.name_en.lower(): g for g in Gender.objects.all()}
        constituency_map = {c.ordering: c for c in Constituency.objects.all()}
        existing_ids     = set(MP.objects.values_list('mp_id', flat=True))

        # ── Open workbook ─────────────────────────────────────────────────────
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        def col(row_num, name):
            idx = headers.index(name)
            val = ws.cell(row_num, idx + 1).value
            return str(val).strip() if val is not None else ''

        # ── Stats ─────────────────────────────────────────────────────────────
        total = skipped = created = errors = 0
        error_log = []

        for row_num in range(2, ws.max_row + 1):
            total += 1
            mp_id = col(row_num, 'MP ID')

            # ── Skip if already imported ──────────────────────────────────────
            if skip_existing and mp_id in existing_ids:
                skipped += 1
                self.stdout.write(f'  SKIP  row {row_num} | {mp_id} already exists')
                continue

            # ── Resolve FK fields ─────────────────────────────────────────────
            member_type_raw = col(row_num, 'Member Type')
            member_type     = MEMBER_TYPE_MAP.get(member_type_raw.lower(), 'direct')
            is_reserved     = (member_type == 'reserved')

            religion     = _resolve(religion_map,  col(row_num, 'Religion'),     RELIGION_ALIASES)
            gender       = _resolve(gender_map,    col(row_num, 'Gender'))
            home_dist    = _resolve(district_map,  col(row_num, 'Home District'), DISTRICT_ALIASES)

            division_obj = None
            district_obj = None
            if not is_reserved:
                division_obj = _resolve(division_map, col(row_num, 'Division'), DIVISION_ALIASES)
                district_obj = _resolve(district_map,  col(row_num, 'District'),  DISTRICT_ALIASES)

            constituency_obj = None
            constituency_str = col(row_num, 'Constituency')
            if constituency_str:
                try:
                    seat_num         = int(constituency_str.split()[0])
                    constituency_obj = constituency_map.get(seat_num)
                except (ValueError, IndexError):
                    pass

            dob = _parse_date(col(row_num, 'Date Of Birth(DD/MM/YYYY)'))

            mobile     = _norm_mobile(col(row_num, 'Mobile Number'))
            alt_mobile = _norm_mobile(col(row_num, 'Alternative Mobile Number'))
            whatsapp   = _norm_mobile(col(row_num, 'Whatsapp Number'))
            email      = col(row_num, 'Official Email')

            try:
                times_elected = int(col(row_num, 'How many times have been elected') or 1)
            except ValueError:
                times_elected = 1

            # ── Collect row-level warnings ────────────────────────────────────
            row_warnings = []
            if not constituency_obj:
                row_warnings.append(f'constituency not resolved: "{constituency_str}"')
            if not is_reserved and not division_obj and col(row_num, 'Division'):
                row_warnings.append(f'division not resolved: "{col(row_num, "Division")}"')
            if not is_reserved and not district_obj and col(row_num, 'District'):
                row_warnings.append(f'district not resolved: "{col(row_num, "District")}"')
            if not home_dist and col(row_num, 'Home District'):
                row_warnings.append(f'home_district not resolved: "{col(row_num, "Home District")}"')
            if not religion and col(row_num, 'Religion'):
                row_warnings.append(f'religion not resolved: "{col(row_num, "Religion")}"')
            if not gender and col(row_num, 'Gender'):
                row_warnings.append(f'gender not resolved: "{col(row_num, "Gender")}"')
            if not dob and col(row_num, 'Date Of Birth(DD/MM/YYYY)'):
                row_warnings.append(f'dob not parsed: "{col(row_num, "Date Of Birth(DD/MM/YYYY)")}"')

            if row_warnings:
                for w in row_warnings:
                    self.stdout.write(self.style.WARNING(
                        f'  WARN  row {row_num} | {mp_id} | {w}'
                    ))

            # ── Dry-run: report and continue ──────────────────────────────────
            if dry_run:
                name_en = col(row_num, 'MP Name(English)')
                self.stdout.write(
                    f'  OK    row {row_num} | {mp_id} | {name_en} | '
                    f'{member_type} | constituency={constituency_obj} | '
                    f'division={division_obj} | district={district_obj}'
                )
                created += 1
                continue

            # ── Save ──────────────────────────────────────────────────────────
            try:
                with transaction.atomic():
                    mp = MP.objects.create(
                        mp_id       = mp_id,
                        parliament  = parliament,
                        member_type = member_type,
                        name_bn     = col(row_num, 'MP Name(Bangla)'),
                        name_en     = col(row_num, 'MP Name(English)'),
                        nid         = col(row_num, 'Nid') or None,
                        dob         = dob,
                        religion    = religion,
                        gender      = gender,
                        home_district = home_dist,
                    )

                    ElectionInfo.objects.create(
                        mp            = mp,
                        parliament    = parliament,
                        constituency  = constituency_obj,
                        times_elected = times_elected,
                    )

                    Address.objects.create(
                        mp           = mp,
                        address_type = 'present',
                        division     = division_obj,
                        district     = district_obj,
                        mobile       = mobile,
                        alt_mobile   = alt_mobile,
                        whatsapp     = whatsapp,
                        email        = email,
                    )

                existing_ids.add(mp_id)
                created += 1
                self.stdout.write(
                    self.style.SUCCESS(f'  SAVED row {row_num} | {mp_id} | {mp.name_en}')
                )

            except Exception as exc:
                errors += 1
                msg = f'row {row_num} | {mp_id} | {exc}'
                error_log.append(msg)
                self.stdout.write(self.style.ERROR(f'  ERROR {msg}'))

        # ── Summary ───────────────────────────────────────────────────────────
        self.stdout.write('')
        self.stdout.write('─' * 60)
        self.stdout.write(f'Total rows   : {total}')
        self.stdout.write(f'Skipped      : {skipped}  (already in DB)')
        action = 'Would create' if dry_run else 'Created'
        self.stdout.write(f'{action}     : {created}')
        if not dry_run:
            self.stdout.write(f'Errors       : {errors}')
        self.stdout.write('─' * 60)

        if error_log:
            self.stdout.write(self.style.ERROR('\nFailed rows:'))
            for msg in error_log:
                self.stdout.write(self.style.ERROR(f'  {msg}'))

        if dry_run:
            self.stdout.write(self.style.WARNING(
                '\nDry run complete. Run without --dry-run to import.'
            ))
        elif errors == 0:
            self.stdout.write(self.style.SUCCESS('\nImport complete — no errors.'))
        else:
            self.stdout.write(self.style.WARNING(
                f'\nImport complete with {errors} error(s). Fix and re-run.'
            ))
